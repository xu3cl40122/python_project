#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from os import listdir
from os.path import isfile, isdir, join
import re


def findImg(re_rule, files):
    i = 0
    isFound = False
    while i < len(files):
        isFound = False
        match_object = re.match(re_rule, files[i], re.IGNORECASE)
        if match_object != None:
               isFound = True
               break
        i += 1
    if isFound == False:
        print('------------' + re_rule + ' not found -------------')
        return False
    return True


def writeOutput(isFound, row, tr_id, imgToFind):
    if isFound:
        ws_w.cell(row=row, column=3).value = 'o'
    else:
        ws_w.cell(row=row, column=3).value = 'not found'
    ws_w.cell(row=row, column=2).value = imgToFind
    ws_w.cell(row=row, column=1).value = tr_id





# --- main ---  
def main(subFile='/05', line=1228, outputFile='output_2.xlsx', input_file='./tr_pic.xlsx'):
    global wb_write, wb_read , ws_w, ws_r
    # init excel to write
    wb_write = Workbook()
    ws_w = wb_write.active
    ws_w.title = "output"
    # init excel to read
    wb_read = load_workbook(filename=input_file)
    ws_r = wb_read['ts1']
    tr_id = '0'

    # --- 主迴圈 依據 excel 的 row 跑 ---
    for r in range(1, line):
        # 確認搜尋條件是否合理
        check_id = str(ws_r.cell(row=r, column=2).value)
        check_pic_name = str(ws_r.cell(row=r, column=3).value)
        if check_id == 'None' or check_id == "":
            ws_w.cell(row=r, column=3).value = 'null line'
            continue
        if (len(check_pic_name) < 3) or check_pic_name == 'None':
            ws_w.cell(row=r, column=3).value = 'target??'
            continue
        
        #判斷是否換一條步道
        if(tr_id != str(ws_r.cell(row=r, column=2).value)):
            tr_id = str(ws_r.cell(row=r, column=2).value)
            print('change'+tr_id)
            # 依 loop 更換路徑
            mypath = "./" + tr_id + subFile
            # 取得路徑下所有檔案與子目錄名稱
            try:
                files = listdir(mypath)
            except FileNotFoundError:
                ws_w.cell(row=r, column=3).value = 'error'
                print('000000 error 000000')
                print(mypath)
                continue
            imgToFind = str(ws_r.cell(row=r, column=3).value)
            p_isFound = findImg(imgToFind, files)
            writeOutput(p_isFound, r, tr_id, imgToFind)

        else:
            imgToFind = str(ws_r.cell(row=r, column=3).value)
            p_isFound = findImg(imgToFind, files)
            writeOutput(p_isFound, r, tr_id, imgToFind)
    print('save as : '+ outputFile)
    wb_write.save(outputFile)


main('/05', 121, 'output_test.xlsx', './test.xlsx')
